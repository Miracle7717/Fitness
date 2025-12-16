from django.http import HttpResponse
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from openpyxl import Workbook
import io

@login_required
def export_clients_pdf(request):
    """Экспорт клиентов в PDF"""
    from clients.models import Client
    
    # Создаем PDF
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    
    # Заголовок
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, 800, "Список клиентов фитнес-клуба")
    
    # Данные
    p.setFont("Helvetica", 10)
    y = 750
    clients = Client.objects.all()
    
    for client in clients:
        if y < 50:
            p.showPage()
            p.setFont("Helvetica", 10)
            y = 750
        
        p.drawString(50, y, f"{client.get_full_name()}")
        p.drawString(200, y, f"{client.phone}")
        p.drawString(300, y, f"{client.email or '-'}")
        y -= 20
    
    p.save()
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="clients.pdf"'
    return response

@login_required
def export_payments_excel(request):
    """Экспорт платежей в Excel"""
    from payments.models import Payment
    
    # Создаем Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Платежи"
    
    # Заголовки
    headers = ['ID', 'Дата', 'Клиент', 'Сумма', 'Тип', 'Метод', 'Статус']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Данные
    payments = Payment.objects.all().select_related('client')
    for row_num, payment in enumerate(payments, 2):
        ws.cell(row=row_num, column=1, value=payment.id)
        ws.cell(row=row_num, column=2, value=payment.payment_date.strftime("%d.%m.%Y %H:%M"))
        ws.cell(row=row_num, column=3, value=payment.client.get_full_name())
        ws.cell(row=row_num, column=4, value=float(payment.amount))
        ws.cell(row=row_num, column=5, value=payment.get_payment_type_display_name())
        ws.cell(row=row_num, column=6, value=payment.get_payment_method_display_name())
        ws.cell(row=row_num, column=7, value=payment.get_status_display_name())
    
    # Сохраняем
    buffer = io.BytesIO()
    wb.save(buffer)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="payments.xlsx"'
    return response