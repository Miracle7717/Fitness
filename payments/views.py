from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.utils import timezone
import datetime
from .models import Payment, Reminder
from .forms import PaymentForm, PaymentSearchForm, ReminderForm
from clients.models import Client
from subscriptions.models import Membership

@login_required
def payment_list(request):
    """Список всех платежей"""
    form = PaymentSearchForm(request.GET or None)
    payments = Payment.objects.all().select_related('client', 'membership', 'membership_plan').order_by('-payment_date')
    
   
    if form.is_valid():
        search = form.cleaned_data.get('search')
        status = form.cleaned_data.get('status')
        payment_type = form.cleaned_data.get('payment_type')
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')
        
        if search:
            payments = payments.filter(
                Q(client__first_name__icontains=search) |
                Q(client__last_name__icontains=search) |
                Q(client__phone__icontains=search)
            )
        
        if status:
            payments = payments.filter(status=status)
        
        if payment_type:
            payments = payments.filter(payment_type=payment_type)
        
        if start_date:
            payments = payments.filter(payment_date__date__gte=start_date)
        
        if end_date:
            payments = payments.filter(payment_date__date__lte=end_date)
    

    total_amount = payments.filter(status='completed').aggregate(total=Sum('amount'))['total'] or 0
    total_count = payments.count()
    completed_count = payments.filter(status='completed').count()
    

    paginator = Paginator(payments, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'form': form,
        'total_amount': total_amount,
        'total_count': total_count,
        'completed_count': completed_count,
    }
    return render(request, 'payments/payment_list.html', context)

@login_required
def payment_detail(request, pk):
    """Детальная информация о платеже"""
    payment = get_object_or_404(Payment, pk=pk)
    
    context = {
        'payment': payment,
    }
    return render(request, 'payments/payment_detail.html', context)

@login_required
def payment_create(request):
    """Создание нового платежа"""
    if request.method == 'POST':
        form = PaymentForm(request.POST, request.FILES)
        if form.is_valid():
            payment = form.save()
            messages.success(request, f'Платеж на сумму {payment.amount} руб. создан!')
            
            if payment.status == 'completed' and payment.membership:

                reminder_date = payment.period_end - datetime.timedelta(days=7)
                if reminder_date > timezone.now().date():
                    Reminder.objects.create(
                        client=payment.client,
                        membership=payment.membership,
                        payment=payment,
                        reminder_type='subscription_expiry',
                        send_date=reminder_date,
                        send_method='email',
                        subject=f'Напоминание об истечении абонемента',
                        message=f'Уважаемый {payment.client.get_full_name()}, ваш абонемент истекает {payment.period_end}.'
                    )
            
            return redirect('payment_detail', pk=payment.pk)
    else:
        form = PaymentForm()
    
    context = {'form': form}
    return render(request, 'payments/payment_form.html', context)

@login_required
def payment_update(request, pk):
    """Редактирование платежа"""
    payment = get_object_or_404(Payment, pk=pk)
    
    if request.method == 'POST':
        form = PaymentForm(request.POST, request.FILES, instance=payment)
        if form.is_valid():
            payment = form.save()
            messages.success(request, 'Платеж обновлен!')
            return redirect('payment_detail', pk=payment.pk)
    else:
        form = PaymentForm(instance=payment)
    
    context = {
        'form': form,
        'payment': payment,
    }
    return render(request, 'payments/payment_form.html', context)

@login_required
def payment_delete(request, pk):
    """Удаление платежа"""
    payment = get_object_or_404(Payment, pk=pk)
    
    if request.method == 'POST':
        amount = payment.amount
        payment.delete()
        messages.success(request, f'Платеж на сумму {amount} руб. удален!')
        return redirect('payment_list')
    
    context = {'payment': payment}
    return render(request, 'payments/payment_confirm_delete.html', context)

@login_required
def payment_statistics(request):
    """Статистика по платежам"""
    today = timezone.now().date()
    
    total_payments = Payment.objects.filter(status='completed').aggregate(total=Sum('amount'))['total'] or 0
    total_count = Payment.objects.count()
    

    from django.db.models.functions import TruncMonth
    monthly_stats = Payment.objects.filter(status='completed').annotate(
        month=TruncMonth('payment_date')
    ).values('month').annotate(
        total=Sum('amount'),
        count=models.Count('id')
    ).order_by('-month')[:12]
    

    type_stats = Payment.objects.filter(status='completed').values(
        'payment_type'
    ).annotate(
        total=Sum('amount'),
        count=models.Count('id')
    ).order_by('-total')
    

    method_stats = Payment.objects.filter(status='completed').values(
        'payment_method'
    ).annotate(
        total=Sum('amount'),
        count=models.Count('id')
    ).order_by('-total')
    
    today_payments = Payment.objects.filter(
        status='completed',
        payment_date__date=today
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    month_start = today.replace(day=1)
    month_payments = Payment.objects.filter(
        status='completed',
        payment_date__date__gte=month_start
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    context = {
        'total_payments': total_payments,
        'total_count': total_count,
        'monthly_stats': monthly_stats,
        'type_stats': type_stats,
        'method_stats': method_stats,
        'today_payments': today_payments,
        'month_payments': month_payments,
        'today': today,
    }
    return render(request, 'payments/payment_statistics.html', context)

@login_required
def reminder_list(request):
    """Список напоминаний"""
    reminders = Reminder.objects.all().select_related('client', 'membership').order_by('send_date')
    
    pending_reminders = reminders.filter(send_status='pending')
    sent_reminders = reminders.filter(send_status='sent')
    failed_reminders = reminders.filter(send_status='failed')
    

    overdue_reminders = pending_reminders.filter(send_date__lt=timezone.now())
    
    context = {
        'pending_reminders': pending_reminders,
        'sent_reminders': sent_reminders,
        'failed_reminders': failed_reminders,
        'overdue_reminders': overdue_reminders,
    }
    return render(request, 'payments/reminder_list.html', context)

@login_required
def reminder_create(request):
    """Создание напоминания"""
    if request.method == 'POST':
        form = ReminderForm(request.POST)
        if form.is_valid():
            reminder = form.save()
            messages.success(request, 'Напоминание создано!')
            return redirect('reminder_list')
    else:
        form = ReminderForm()
    
    context = {'form': form}
    return render(request, 'payments/reminder_form.html', context)

@login_required
def reminder_send_now(request, pk):
    """Отправка напоминания сейчас"""
    reminder = get_object_or_404(Reminder, pk=pk)
    
    reminder.mark_as_sent()
    messages.success(request, 'Напоминание отправлено!')
    
    return redirect('reminder_list')

@login_required
def debtors_list(request):
    """Список должников"""

    today = timezone.now().date()
    expired_memberships = Membership.objects.filter(
        status='active',
        end_date__lt=today
    ).select_related('client')
    

    debtors = {}
    for membership in expired_memberships:
        client = membership.client
        if client not in debtors:
            debtors[client] = []
        debtors[client].append(membership)
    
 
    sorted_debtors = sorted(
        debtors.items(),
        key=lambda x: max((today - m.end_date).days for m in x[1]),
        reverse=True
    )
    
    context = {
        'debtors': sorted_debtors,
        'today': today,
    }
    return render(request, 'payments/debtors_list.html', context)

@login_required
def create_subscription_payment(request, membership_id):
    """Быстрое создание платежа за абонемент"""
    membership = get_object_or_404(Membership, pk=membership_id)
    
    if request.method == 'POST':
        form = PaymentForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.client = membership.client
            payment.membership = membership
            payment.membership_plan = membership.plan
            payment.amount = membership.plan.price
            payment.payment_type = 'subscription'
            payment.save()
            
            messages.success(request, f'Платеж за абонемент создан!')
            return redirect('payment_detail', pk=payment.pk)
    else:
    
        initial_data = {
            'client': membership.client,
            'membership': membership,
            'membership_plan': membership.plan,
            'amount': membership.plan.price,
            'payment_type': 'subscription',
            'payment_date': timezone.now(),
            'period_start': membership.start_date,
            'period_end': membership.end_date,
        }
        form = PaymentForm(initial=initial_data)
    
    context = {
        'form': form,
        'membership': membership,
    }
    return render(request, 'payments/create_subscription_payment.html', context)

@login_required
def export_payments_excel(request):
    """Экспорт платежей в Excel"""
    from django.http import HttpResponse
    import io
    from openpyxl import Workbook
    
    # Создаем Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Платежи"
    
    # Заголовки
    headers = ['ID', 'Дата', 'Клиент', 'Сумма', 'Тип', 'Метод', 'Статус']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Данные
    payments = Payment.objects.all().select_related('client')
    for row_num, payment in enumerate(payments, 2):
        ws.cell(row=row_num, column=1, value=payment.id)
        ws.cell(row=row_num, column=2, value=payment.payment_date.strftime("%d.%m.%Y %H:%M"))
        ws.cell(row=row_num, column=3, value=payment.client.get_full_name())
        ws.cell(row=row_num, column=4, value=float(payment.amount))
        ws.cell(row=row_num, column=5, value=payment.get_payment_type_display_name())
        ws.cell(row=row_num, column=6, value=payment.get_payment_method_display_name())
        ws.cell(row=row_num, column=7, value=payment.get_status_display_name())
    
    # Сохраняем
    buffer = io.BytesIO()
    wb.save(buffer)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="payments.xlsx"'
    return response